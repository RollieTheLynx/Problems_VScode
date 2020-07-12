# -*- coding: utf-8 -*-
"""
Created on Thu Jun 11 20:59:22 2020

@author: Rollie
"""
import requests
from bs4 import BeautifulSoup
import pandas
import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

link = "http://elvpr.ru/ru/catalog/igbt-modules/"
dataframe = pandas.DataFrame({})
headers = {"User-Agent": "Mozilla/5.0"}
response = requests.get(link, headers = headers)
# print(response.status_code)

soup = BeautifulSoup(response.content, "html.parser")

table = soup.find("div", class_="product-table")  # находим таблицу

header_row = table.find("div", class_="product-table__header product-table__row")  # находим в ней заголовок
column_names = header_row.find_all("div", class_="product-table__el")
headers = []
for title in column_names:
    headers.append(title.get_text())  # сохраняем заголовки столбцов в список

table_rows = table.find_all("a", class_="product-table__row")  # находим в ней строки приборов
devices = []  # список с подсписками для каждого прибора
for row in table_rows:
    new_device = []
    row_elements = row.find_all("div", class_="product-table__el")
    for element in row_elements:
        new_device.append(element.get_text())
    new_device.append(row["href"])
    devices.append(new_device)


# вносим в pandas dataframe
names = [item[0] for item in devices]  # списки для столбцов pandas
vces = [item[1] for item in devices]
ics = [item[2] for item in devices]
vcesats = [item[3] for item in devices]
vfs = [item[4] for item in devices]
tjmaxs = [item[5] for item in devices]
visols = [item[6] for item in devices]
sizes = [item[7] for item in devices]
housings = [item[8] for item in devices]
schemes = [item[9] for item in devices]
uses = [item[10] for item in devices]
status = [item[11] for item in devices]
links = [item[13] for item in devices]

dataframe = pandas.DataFrame({
    headers[0]: names,
    headers[1]: vces,
    headers[2]: ics,
    headers[3]: vcesats,
    headers[4]: vfs,
    headers[5]: tjmaxs,
    headers[6]: visols,
    headers[7]: sizes,
    headers[8]: housings,
    headers[9]: schemes,
    headers[10]: uses,
    headers[11]: status,
    "Ссылка": links
})
dataframe = dataframe.append(dataframe)

# запись dataframe в excel
try:
    wb = openpyxl.load_workbook("Elvpr IGBTs.xlsx")
    ws = wb.create_sheet()
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active

now = datetime.datetime.now()
ws.title = "IGBT " + now.strftime("%Y-%m-%d %H-%M-%S")

for r in dataframe_to_rows(dataframe, index=False, header=True):
    ws.append(r)

wb.save('Elvpr IGBTs.xlsx')