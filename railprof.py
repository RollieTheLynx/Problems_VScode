# -*- coding: utf-8 -*-
"""
Created on Sat May 23 22:40:54 2020

@author: Rollie
"""
import requests
from bs4 import BeautifulSoup
import pandas
from time import sleep
from random import randrange
import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import brotli


url = "https://www.railpro.co.uk/business-directory"
headers = {
    "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Encoding":"gzip, deflate, br",
    "Accept-Language":"ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3",
    "Connection":"keep-alive",
    "Cookie":"__cfduid=d0bdb4a215cc527d3a5d93e808330be631615619521; _ga=GA1.3.2035340857.1615619488; _gid=GA1.3.548665206.1615619488; wpjb_transient_id=1615619525-9263; strack_tracking_code=3596634.43f1ec417b2faffef76c558e18c58396",
    "Host":"www.railpro.co.uk",
    "Referer":"https://www.railpro.co.uk/business-directory",
    "Upgrade-Insecure-Requests":'1',
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:86.0) Gecko/20100101 Firefox/86.0"
    }

response = requests.get(url, headers = headers)
page_html = brotli.decompress(soup)
print(page_html)
soup = BeautifulSoup(response.content, "html.parser")
print(soup)


from urllib.request import urlopen as uReq

my_url = "https://www.railpro.co.uk/business-directory"

uClient = uReq(my_url, headers=headers)
page_html = uClient.read()
uClient.close()

print(uClient.headers['Content-Encoding']) # `br` means `brotli`










#%%
def scrapePage(link, dataframe):
    response = requests.get(link)
    #print(response.status_code)

    soup = BeautifulSoup(response.content, "html.parser")
   
    table = soup.find("table", class_ = "views-table cols-6") #находим таблицу
    table_rows = table.find_all("tr") #находим в ней строки
    
    headers = [] # список с заголовками из th
    t_head = table_rows[0].find_all("th")
    for x in range(0,len(t_head)):
        headers.append(t_head[x].get_text())
    
    devices = [] #список с подсписками для каждого прибора
    t_body = soup.find_all("tr") #строки тела таблицы
    for line in range(1,len(t_body)): #в каждой строке, кроме заголовка...
        new_device = [] 
        new_row = t_body[line].find_all("td")
        for y in range(0,len(new_row)):
            new_device.append(new_row[y].get_text()) #...составляем список new_device из значений (new_row) строки
        devices.append(new_device) #добавляем new_device в devices
        
# вносим в pandas dataframe
    names = [item[0] for item in devices] #списки для столбцов pandas
    vces = [item[1] for item in devices]
    ics = [item[2] for item in devices]
    vcesats = [item[3] for item in devices]
    ips = [item[4] for item in devices]
    topos = [item[5] for item in devices]
    
    dataframe2 = pandas.DataFrame({
        headers[0]: names,
        headers[1]: vces,
        headers[2]: ics,
        headers[3]: vcesats,
        headers[4]: ips,
        headers[5]: topos
        })
    dataframe = dataframe.append(dataframe2)
    
    button = soup.find("a", {"title":"На следующую страницу"})
    try:
        new_link = "https://www.angstrem.ru" + button['href']
    except:
        new_link = "0"
    
    return new_link, dataframe
#%%    
link = "https://www.railpro.co.uk/business-directory"
dataframe = pandas.DataFrame({})

while True:
    new_link, new_dataframe = scrapePage(link, dataframe)
    if new_link == "0":
        break
    else:
        link = new_link
        dataframe = new_dataframe
        sleep(randrange(3,15))
        
#запись dataframe в excel
try:
    wb = openpyxl.load_workbook("Angstrem v7.xlsx")
    ws = wb.create_sheet()
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    
now = datetime.datetime.now()
ws.title = "IGBT " + now.strftime("%Y-%m-%d %H-%M-%S")
    
for r in dataframe_to_rows(new_dataframe, index=False, header=True):
    ws.append(r)    
            
wb.save('Angstrem v7.xlsx')
