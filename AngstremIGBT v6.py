# -*- coding: utf-8 -*-
"""
Created on Sat May 23 22:40:54 2020

@author: Rollie
"""
import requests
from bs4 import BeautifulSoup
import pandas
from time import sleep
from random import randrange
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

#%%
def scrapePage(link, dataframe):
    response = requests.get(link)
    #print(response.status_code)

    soup = BeautifulSoup(response.content, "html.parser")
   
    table = soup.find("table", class_ = "views-table cols-6") #находим таблицу
    table_rows = table.find_all("tr") #находим в ней строки
    
    headers = [] # список с заголовками из th
    t_head = table_rows[0].find_all("th")
    for x in range(0,len(t_head)):
        headers.append(t_head[x].get_text())
    
    devices = [] #список с подсписками для каждого прибора
    t_body = soup.find_all("tr") #строки тела таблицы
    for line in range(1,len(t_body)): #в каждой строке, кроме заголовка...
        new_device = [] 
        new_row = t_body[line].find_all("td")
        for y in range(0,len(new_row)):
            new_device.append(new_row[y].get_text()) #...составляем список new_device из значений (new_row) строки
        devices.append(new_device) #добавляем new_device в devices
        
# вносим в pandas dataframe
    names = [item[0] for item in devices] #списки для столбцов pandas
    vces = [item[1] for item in devices]
    ics = [item[2] for item in devices]
    vcesats = [item[3] for item in devices]
    ips = [item[4] for item in devices]
    topos = [item[5] for item in devices]
    
    dataframe2 = pandas.DataFrame({
        headers[0]: names,
        headers[1]: vces,
        headers[2]: ics,
        headers[3]: vcesats,
        headers[4]: ips,
        headers[5]: topos
        })
    dataframe = dataframe.append(dataframe2)
    
    button = soup.find("a", {"title":"На следующую страницу"})
    try:
        new_link = "https://www.angstrem.ru" + button['href']
    except:
        new_link = "0"
    
    return new_link, dataframe
    
#%%
link = "https://www.angstrem.ru/ru/catalog/silovye-poluprovodnikovye-pribory/igbt-moduli"
dataframe = pandas.DataFrame({})

while True:
    new_link, new_dataframe = scrapePage(link, dataframe)
    if new_link == "0":
        break
    else:
        link = new_link
        dataframe = new_dataframe
        sleep(randrange(3,15))
        
#запись dataframe в excel
now = datetime.datetime.now()

try:
    writer = pandas.ExcelWriter('Angstrem v6.xlsx', engine='openpyxl', mode='a')
except FileNotFoundError:
    writer = pandas.ExcelWriter('Angstrem v6.xlsx', engine='xlsxwriter')
    writer.save()
    writer = pandas.ExcelWriter('Angstrem v6.xlsx', engine='openpyxl', mode='a')
new_dataframe.to_excel(writer, index=False, header=True, sheet_name="IGBT " + now.strftime("%Y-%m-%d %H-%M-%S"))

writer.save()


#%%
        try:
            wb = openpyxl.load_workbook("Angstrem v6.xlsx")
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            
        for r in dataframe_to_rows(new_dataframe, index=False, header=True):
            ws.append(r)    
            
            
            
            
            
            wb.save('orders_export.xlsx')
#%%
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
