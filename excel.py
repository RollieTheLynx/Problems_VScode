# -*- coding: utf-8 -*-
"""
Created on Fri May  8 22:47:47 2020

@author: Rollie
"""


import openpyxl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Sheet1']
    
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,3)
        correctedprice = (cell.value)*0.9
        correctedcell = sheet.cell(row,4)
        correctedcell.value = correctedprice
    
    values = Reference(sheet, 
              min_row=2, 
              max_row=sheet.max_row,
              min_col=4,
              max_col=4) 
    
    chart = BarChart()   
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)